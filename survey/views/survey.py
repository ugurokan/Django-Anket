from django.contrib.auth.models import User
from django.db import transaction
from django.http import Http404, HttpResponse
from django.contrib.auth.decorators import login_required
from django.forms.formsets import formset_factory
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.contrib.auth import logout
from openpyxl.workbook import Workbook

from ..models import Survey, Question, Answer, Submission
from ..forms import SurveyForm, QuestionForm, OptionForm, AnswerForm, BaseAnswerFormSet

from django.contrib.auth import REDIRECT_FIELD_NAME
from django.contrib.admin.views.decorators import user_passes_test


def superuser_required(view_func=None, redirect_field_name=REDIRECT_FIELD_NAME,
                       login_url='/login/'):
    """
    Decorator for views that checks that the user is logged in and is a
    superuser, redirecting to the login page if necessary.
    """
    actual_decorator = user_passes_test(
        lambda u: u.is_active and u.is_superuser,
        login_url=login_url,
        redirect_field_name=redirect_field_name
    )
    if view_func:
        return actual_decorator(view_func)
    return actual_decorator


def download_excel(request, pk):
    survey = Survey.objects.get(id=pk)
    excel_file = export_survey_responses_to_excel(pk)
    print(f"Excel file '{excel_file}' containing survey responses has been created.")
    # Path to the existing Excel file
    excel_file_path = f'{survey.title}_responses.xlsx'

    # Open the Excel file in binary mode
    with open(excel_file_path, 'rb') as excel_file:
        # Read the contents of the file
        excel_content = excel_file.read()

    # Create an HTTP response with the Excel file as attachment
    response = HttpResponse(excel_content,
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Cevaplar.xlsx"'
    return response


@login_required
def survey_list(request):
    """User can view all their surveys"""
    surveys = Survey.objects.filter(creator=request.user).order_by("-created_at").all()
    # Todo kullanıcı 3 ü çözdümü
    if not request.user.is_superuser:
        survey3 = get_object_or_404(Survey, pk=3)
        survey2 = get_object_or_404(Survey, pk=2)
        survey1 = get_object_or_404(Survey, pk=1)

        user_answers3 = Answer.objects.filter(submission__survey=survey3, user=request.user)
        user_answers2 = Answer.objects.filter(submission__survey=survey2, user=request.user)
        user_answers1 = Answer.objects.filter(submission__survey=survey1, user=request.user)

        if not len(user_answers3) == 0:
            return render(request, "survey/thanks.html", {"survey": survey3})
        elif not len(user_answers2) == 0:
            return redirect("survey-start", pk=2)
        elif len(user_answers1) == 0:
            return redirect("survey-start", pk=1)
        elif len(user_answers1) > 0:
            return redirect("survey-slayt", pk=1)
    return render(request, "survey/list.html", {"surveys": surveys})


@login_required
def detail(request, pk):
    """User can view an active survey"""
    try:
        survey = Survey.objects.prefetch_related("question_set__option_set").get(
            pk=pk, creator=request.user, is_active=True
        )
    except Survey.DoesNotExist:
        raise Http404()

    questions = survey.question_set.all()

    # Calculate the results.
    # This is a naive implementation and it could be optimised to hit the database less.
    # See here for more info on how you might improve this code: https://docs.djangoproject.com/en/3.1/topics/db/aggregation/

    for question in questions:
        option_pks = question.option_set.values_list("pk", flat=True)
        total_answers = Answer.objects.filter(option_id__in=option_pks).count()
        for option in question.option_set.all():
            num_answers = Answer.objects.filter(option=option).count()
            option.percent = 100.0 * num_answers / total_answers if total_answers else 0

    host = request.get_host()
    public_path = reverse("survey-start", args=[pk])
    public_url = f"{request.scheme}://{host}{public_path}"
    num_submissions = survey.submission_set.filter(is_complete=True).count()
    return render(
        request,
        "survey/detail.html",
        {
            "survey": survey,
            "public_url": public_url,
            "questions": questions,
            "num_submissions": num_submissions,
        },
    )


@superuser_required
@login_required
def create(request):
    """User can create a new survey"""
    if request.method == "POST":
        form = SurveyForm(request.POST)
        if form.is_valid():
            survey = form.save(commit=False)
            survey.creator = request.user
            survey.save()
            return redirect("survey-edit", pk=survey.id)
    else:
        form = SurveyForm()

    return render(request, "survey/create.html", {"form": form})


@login_required
def delete(request, pk):
    """User can delete an existing survey"""
    survey = get_object_or_404(Survey, pk=pk, creator=request.user)
    if request.method == "POST":
        survey.delete()

    return redirect("survey-list")


@login_required
def edit(request, pk):
    """User can add questions to a draft survey, then acitvate the survey"""
    try:
        survey = Survey.objects.prefetch_related("question_set__option_set").get(
            pk=pk, creator=request.user, is_active=False
        )
    except Survey.DoesNotExist:
        raise Http404()

    if request.method == "POST":
        survey.is_active = True
        survey.save()
        return redirect("survey-detail", pk=pk)
    else:
        questions = survey.question_set.all()
        return render(request, "survey/edit.html", {"survey": survey, "questions": questions})


@login_required
def question_create(request, pk):
    """User can add a question to a draft survey"""
    survey = get_object_or_404(Survey, pk=pk, creator=request.user)
    if request.method == "POST":
        form = QuestionForm(request.POST)
        if form.is_valid():
            question = form.save(commit=False)
            question.survey = survey
            question.save()
            return redirect("survey-option-create", survey_pk=pk, question_pk=question.pk)
    else:
        form = QuestionForm()

    return render(request, "survey/question.html", {"survey": survey, "form": form})


@login_required
def option_create(request, survey_pk, question_pk):
    """User can add options to a survey question"""
    survey = get_object_or_404(Survey, pk=survey_pk, creator=request.user)
    question = Question.objects.get(pk=question_pk)
    if request.method == "POST":
        form = OptionForm(request.POST)
        if form.is_valid():
            option = form.save(commit=False)
            option.question_id = question_pk
            option.save()
    else:
        form = OptionForm()

    options = question.option_set.all()
    return render(
        request,
        "survey/options.html",
        {"survey": survey, "question": question, "options": options, "form": form},
    )


@login_required
def start(request, pk):
    """Survey-taker can start a survey"""
    survey = get_object_or_404(Survey, pk=pk, is_active=True)
    if request.method == "POST":
        sub = Submission.objects.create(survey=survey)
        return redirect("survey-submit", survey_pk=pk, sub_pk=sub.pk)

    if pk==3:
        return render(request, "survey/memnuniyet_start.html", {"survey": survey})

    return render(request, "survey/start.html", {"survey": survey})

def submit(request, survey_pk, sub_pk):
    """Survey-taker submit their completed survey."""
    try:
        survey = Survey.objects.prefetch_related("question_set__option_set").get(
            pk=survey_pk, is_active=True
        )
    except Survey.DoesNotExist:
        raise Http404()

    try:
        sub = survey.submission_set.get(pk=sub_pk, is_complete=False)
    except Submission.DoesNotExist:
        raise Http404()

    questions = survey.question_set.all()
    # options = [q.option_set.all() for q in questions]
    options = [q.option_set.all().order_by('id') for q in questions]
    form_kwargs = {"empty_permitted": False, "options": options}
    AnswerFormSet = formset_factory(AnswerForm, extra=len(questions), formset=BaseAnswerFormSet)
    if request.method == "POST":
        formset = AnswerFormSet(request.POST, form_kwargs=form_kwargs)
        if formset.is_valid():
            with transaction.atomic():
                for form in formset:
                    Answer.objects.create(
                        option_id=form.cleaned_data["option"], submission_id=sub_pk, user=request.user
                    )

                sub.is_complete = True
                sub.save()
            if survey_pk == 3:
                print("deneme1")
                return redirect("survey-thanks", pk=survey_pk)
            elif survey_pk == 2:
                return redirect("survey-start", pk=3)
                print("deneme2")

            elif survey_pk == 1:
                print("deneme3")
                return redirect("survey-slayt", pk=survey_pk)

    else:
        formset = AnswerFormSet(form_kwargs=form_kwargs)

    print(form_kwargs)
    question_forms = zip(questions, formset)


    return render(
        request,
        "survey/submit.html",
        {"survey": survey, "question_forms": question_forms, "formset": formset},
    )


def thanks(request, pk):
    """Survey-taker receives a thank-you message."""
    survey = get_object_or_404(Survey, pk=pk, is_active=True)
    return render(request, "survey/thanks.html", {"survey": survey})


@login_required
def slayt(request, pk):
    """Survey-taker receives a thank-you message."""
    survey = get_object_or_404(Survey, pk=pk, is_active=True)

    # Get the survey object with the given ID
    survey = get_object_or_404(Survey, pk=1)

    # Get all answers submitted by the current user for the given survey
    user_answers = Answer.objects.filter(submission__survey=survey, user=request.user)

    print(user_answers)
    yanlislar = []
    for answer in user_answers:
        if (answer.option.Dogru_cevap == False):
            print(f"Question: {answer.option.question.prompt}")
            print(f"Answer: {answer.option.text}")
            print(f"Answer: {answer.option.Dogru_cevap}")
            print()
            print("")
            yanlislar.append([answer.option.question.prompt, answer.option.question.link])

    survey_id = 1  # Provide the ID of the survey you want to export responses for
    # excel_file = export_survey_responses_to_excel(survey_id)
    # print(f"Excel file '{excel_file}' containing survey responses has been created.")

    answers = Answer.objects.prefetch_related("survey_set__option_set").filter(
        user=request.user,
        submission=6
    )
    print(answers)

    survey_id = 1  # Provide the ID of the survey you want to export responses for

    return render(
        request,
        "survey/slayt.html",
        {
            "yanlislar": yanlislar,

        },
    )
    # # logout(request)
    # return render(request, "survey/slayt.html", {"survey": survey})


def logout_view(request):
    logout(request)
    return redirect('landing')  # Redirect to home page or any other page after logout


def export_survey_responses_to_excel(survey_id):
    # Get the survey
    survey = Survey.objects.get(id=survey_id)

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Create a dictionary to store question prompts and their corresponding column index
    question_column_mapping = {}

    # Write headers
    ws.append([' '])

    # Get all questions for the survey
    questions = Question.objects.filter(survey=survey)

    # Write question prompts as headers
    for idx, question in enumerate(questions, start=2):  # Starting from column 2
        ws.cell(row=1, column=idx, value=question.prompt)
        question_column_mapping[question.id] = idx

    # Get all submissions for the survey
    submissions = Submission.objects.filter(survey=survey,is_complete=1)




    # Iterate over submissions
    for submission in submissions:
        # Get user associated with the submission
        user = submission.answer_set.first().user
        # # Get user's username
        user_username = user.username


        # Write user's username in the first column
        ws.append([user_username])

        # Create a dictionary to store user's answers for each question
        user_answers_dict = {}

        # Iterate over user's answers
        for answer in submission.answer_set.all():
            # Get question and option text
            question_id = answer.option.question.id
            option_text = answer.option.text

            # Store answer for each question
            user_answers_dict[question_id] = option_text

        # Write user's answers in the corresponding columns
        for question_id, column_idx in question_column_mapping.items():
            answer_text = user_answers_dict.get(question_id, '')
            ws.cell(row=ws.max_row, column=column_idx, value=answer_text)

    # Save workbook
    file_name = f"{survey.title}_responses.xlsx"
    wb.save(file_name)

    return file_name
